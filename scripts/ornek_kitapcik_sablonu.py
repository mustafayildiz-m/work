#!/usr/bin/env python3
"""
Karşı tarafa gönderilecek örnek Excel şablonu oluşturur.
5000 kitapçık toplu yüklemesi için veri formatı.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ════════════════════════════════════════════
# SAYFA 1: Ana Veri Şablonu
# ════════════════════════════════════════════
ws = wb.active
ws.title = "Kitapçık Verileri"

headers = [
    ("kitap_adi", "İlişkili Kitap Adı *", 30,
     "Bu kitapçık hangi kitaba ait? Sistemde kayıtlı kitap adıyla birebir aynı olmalı."),
    ("dil", "Dil Kodu *", 12,
     "Dil kodu: tr, en, ar, de, fr, ru, es, pt, it, zh, ja, ko, ky ..."),
    ("baslik", "Kitapçık Başlığı *", 40,
     "Kitapçığın başlığı (her dil için ayrı satır)."),
    ("icerik", "İçerik (Tam Metin) *", 60,
     "Kitapçığın tam içeriği. Düz metin veya HTML olabilir."),
    ("ozet", "Özet", 40,
     "Kısa özet / tanıtım metni (opsiyonel)."),
    ("yazar", "Yazar", 25,
     "Kitapçık yazarı (opsiyonel)."),
    ("yayin_tarihi", "Yayın Tarihi", 16,
     "YYYY-MM-DD formatında, örn: 2024-06-15 (opsiyonel)."),
    ("siralama", "Sıra No", 10,
     "Aynı kitaptaki kitapçıkların sırası: 1, 2, 3... (opsiyonel)."),
    ("pdf_dosya_adi", "PDF Dosya Adı", 30,
     "Varsa PDF dosyasının adı, örn: 001_giris.pdf — pdfler/ klasöründe olmalı (opsiyonel)."),
    ("kapak_resmi_dosya_adi", "Kapak Resmi Dosya Adı", 30,
     "Varsa kapak resminin adı, örn: 001.jpg — kapaklar/ klasöründe olmalı (opsiyonel)."),
]

# Stiller
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
required_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
optional_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
example_font = Font(color="6B7280", italic=True, size=10)
note_font = Font(color="DC2626", bold=True, size=10)
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB'),
)

# Header satırı
for col_idx, (key, label, width, desc) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Açıklama satırı (2. satır)
for col_idx, (key, label, width, desc) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=desc)
    cell.font = Font(color="6B7280", size=9, italic=True)
    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    cell.border = thin_border

ws.row_dimensions[2].height = 45

# ── Örnek veriler ──
examples = [
    # Tek kitap, tek dil
    ["Kuran Tefsiri", "tr", "1. Bölüm: Fatiha Suresi Tefsiri",
     "Fatiha suresi, Kur'an-ı Kerim'in ilk suresidir. Yedi ayetten oluşur ve namazda okunması farzdır...",
     "Fatiha suresinin detaylı tefsiri ve açıklaması.",
     "Prof. Dr. Ahmet Yılmaz", "2024-01-15", 1, "001_fatiha.pdf", "001_fatiha.jpg"],

    ["Kuran Tefsiri", "tr", "2. Bölüm: Bakara Suresi Tefsiri (1-20)",
     "Bakara suresi Kur'an'ın en uzun suresidir. İlk ayetler iman edenlerin özelliklerini anlatır...",
     "Bakara suresinin ilk 20 ayetinin tefsiri.",
     "Prof. Dr. Ahmet Yılmaz", "2024-01-20", 2, "002_bakara.pdf", "002_bakara.jpg"],

    ["Kuran Tefsiri", "tr", "3. Bölüm: Bakara Suresi Tefsiri (21-40)",
     "Bu bölümde Bakara suresinin 21-40. ayetleri ele alınmaktadır...",
     "Bakara suresinin 21-40. ayetlerinin açıklaması.",
     "Prof. Dr. Ahmet Yılmaz", "2024-02-01", 3, "", ""],

    # Aynı kitap, farklı dil (çoklu dil örneği)
    ["Quran Tafsir", "en", "Chapter 1: Tafsir of Surah Al-Fatiha",
     "Surah Al-Fatiha is the opening chapter of the Quran. It consists of seven verses...",
     "Detailed tafsir of Surah Al-Fatiha.",
     "Prof. Dr. Ahmet Yılmaz", "2024-01-15", 1, "001_fatiha_en.pdf", ""],

    # Farklı kitap
    ["Hadis Koleksiyonu", "tr", "Sahih Buhari - İman Bölümü",
     "İman bölümü, Hz. Peygamber'in iman ile ilgili hadislerini içermektedir...",
     "Buhari'nin iman bölümündeki hadislerin derlemesi.",
     "İmam Buhari", "2024-03-10", 1, "buhari_iman.pdf", "buhari_kapak.jpg"],

    ["Hadis Koleksiyonu", "ar", "صحيح البخاري - كتاب الإيمان",
     "كتاب الإيمان يتضمن أحاديث النبي صلى الله عليه وسلم المتعلقة بالإيمان...",
     "مجموعة أحاديث كتاب الإيمان من صحيح البخاري",
     "الإمام البخاري", "2024-03-10", 1, "buhari_iman_ar.pdf", ""],

    # PDF ve kapak resmi olmayan örnek
    ["Siyer-i Nebi", "tr", "Hz. Muhammed'in Doğumu ve Çocukluğu",
     "Peygamber Efendimiz, Miladi 571 yılında Mekke'de dünyaya gelmiştir...",
     "", "", "2024-05-01", 1, "", ""],

    # Minimal örnek (sadece zorunlu alanlar)
    ["Siyer-i Nebi", "tr", "Hicret Dönemi",
     "Hicret, Hz. Muhammed'in Mekke'den Medine'ye göç etmesidir...",
     "", "", "", 2, "", ""],
]

for row_idx, row_data in enumerate(examples, 3):
    is_required_cols = [True, True, True, True, False, False, False, False, False, False]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = example_font
        cell.fill = required_fill if is_required_cols[col_idx - 1] else optional_fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = thin_border

# Satır yükseklikleri
for r in range(3, 3 + len(examples)):
    ws.row_dimensions[r].height = 50

# Freeze panes
ws.freeze_panes = 'A3'

# ════════════════════════════════════════════
# SAYFA 2: Açıklamalar
# ════════════════════════════════════════════
ws2 = wb.create_sheet("AÇIKLAMALAR - ÖNCELİKLE OKUYUN")

title_font = Font(bold=True, size=14, color="1E40AF")
section_font = Font(bold=True, size=12, color="DC2626")
body_font = Font(size=11)
bold_font = Font(bold=True, size=11)

instructions = [
    ("", title_font),
    ("📋 KİTAPÇIK TOPLU YÜKLEME - VERİ HAZIRLAMA KILAVUZU", title_font),
    ("", body_font),
    ("⚠️  ZORUNLU ALANLAR (* ile işaretli)", section_font),
    ("", body_font),
    ("1. kitap_adi — Kitapçığın ait olduğu kitabın adı. Sistemde kayıtlı olan kitap adıyla BİREBİR AYNI olmalı.", body_font),
    ("2. dil — ISO 639-1 dil kodu kullanın: tr (Türkçe), en (İngilizce), ar (Arapça), de (Almanca)...", body_font),
    ("3. baslik — Kitapçığın başlığı.", body_font),
    ("4. icerik — Kitapçığın tam metin içeriği.", body_font),
    ("", body_font),
    ("📁 DOSYA TESLİM FORMATI", section_font),
    ("", body_font),
    ("Lütfen verileri şu klasör yapısında gönderin:", body_font),
    ("", body_font),
    ("  teslim/", bold_font),
    ("  ├── kitapciklar.xlsx          ← Bu dosyanın doldurulmuş hali", body_font),
    ("  ├── pdfler/                   ← PDF dosyaları (varsa)", body_font),
    ("  │   ├── 001_fatiha.pdf", body_font),
    ("  │   ├── 002_bakara.pdf", body_font),
    ("  │   └── ...", body_font),
    ("  └── kapaklar/                 ← Kapak resimleri (varsa)", body_font),
    ("      ├── 001_fatiha.jpg", body_font),
    ("      ├── 002_bakara.jpg", body_font),
    ("      └── ...", body_font),
    ("", body_font),
    ("🔤 ÇOK DİLLİ KİTAPÇIKLAR", section_font),
    ("", body_font),
    ("Aynı kitapçığın birden fazla dilde çevirisi varsa, HER DİL İÇİN AYRI SATIR açın.", body_font),
    ("Örnek: Fatiha tefsiri hem Türkçe hem İngilizce varsa → 2 satır", body_font),
    ("Sıralama numarasını aynı tutun, böylece eşleştirme yapılabilir.", body_font),
    ("", body_font),
    ("📌 ÖNEMLİ NOTLAR", section_font),
    ("", body_font),
    ("• İçerik düz metin veya HTML olabilir. Paragraflar için satır sonu (Enter) kullanabilirsiniz.", body_font),
    ("• PDF dosya adları, Excel'deki 'pdf_dosya_adi' sütunuyla BİREBİR eşleşmeli.", body_font),
    ("• Kapak resimleri PNG, JPG veya WEBP formatında olmalı.", body_font),
    ("• Tarih formatı: YYYY-MM-DD (örn: 2024-06-15)", body_font),
    ("• Excel'deki 2. satır (sarı) açıklama satırıdır, VERİ GİRMEYİN. Veriler 3. satırdan başlamalı.", body_font),
    ("• Örnek verileri SİLİP kendi verilerinizi girin.", body_font),
    ("", body_font),
    ("❓ CEVAPLANMASI GEREKEN SORULAR", section_font),
    ("", body_font),
    ("Lütfen veriyle birlikte şu bilgileri de iletin:", body_font),
    ("1. Toplamda kaç farklı kitaba ait bu kitapçıklar?", bold_font),
    ("2. Hangi dillerde çevirileri var?", bold_font),
    ("3. PDF dosyaları var mı? Varsa toplam boyut ne kadar?", bold_font),
    ("4. Kapak resimleri var mı?", bold_font),
    ("5. İçerikler düz metin mi yoksa HTML/zengin metin mi?", bold_font),
]

ws2.column_dimensions['A'].width = 100

for row_idx, (text, font) in enumerate(instructions, 1):
    cell = ws2.cell(row=row_idx, column=1, value=text)
    cell.font = font
    cell.alignment = Alignment(wrap_text=True)

# ════════════════════════════════════════════
# SAYFA 3: Desteklenen Dil Kodları
# ════════════════════════════════════════════
ws3 = wb.create_sheet("Dil Kodları")

lang_header_font = Font(bold=True, color="FFFFFF", size=11)
lang_header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

ws3.cell(row=1, column=1, value="Dil Kodu").font = lang_header_font
ws3.cell(row=1, column=1).fill = lang_header_fill
ws3.cell(row=1, column=1).border = thin_border

ws3.cell(row=1, column=2, value="Dil Adı").font = lang_header_font
ws3.cell(row=1, column=2).fill = lang_header_fill
ws3.cell(row=1, column=2).border = thin_border

ws3.cell(row=1, column=3, value="Örnek Ülkeler").font = lang_header_font
ws3.cell(row=1, column=3).fill = lang_header_fill
ws3.cell(row=1, column=3).border = thin_border

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 20
ws3.column_dimensions['C'].width = 40

languages = [
    ("tr", "Türkçe", "Türkiye, Kıbrıs"),
    ("en", "İngilizce", "ABD, İngiltere, Avustralya, Kanada"),
    ("ar", "Arapça", "Suudi Arabistan, Mısır, BAE"),
    ("de", "Almanca", "Almanya, Avusturya, İsviçre"),
    ("fr", "Fransızca", "Fransa, Belçika, Kanada"),
    ("ru", "Rusça", "Rusya, Kazakistan"),
    ("es", "İspanyolca", "İspanya, Meksika, Arjantin"),
    ("pt", "Portekizce", "Brezilya, Portekiz"),
    ("it", "İtalyanca", "İtalya"),
    ("zh", "Çince", "Çin, Tayvan"),
    ("ja", "Japonca", "Japonya"),
    ("ko", "Korece", "Güney Kore"),
    ("ky", "Kırgızca", "Kırgızistan"),
    ("ku", "Kürtçe", "—"),
    ("hi", "Hintçe", "Hindistan"),
    ("ur", "Urduca", "Pakistan"),
    ("id", "Endonezce", "Endonezya"),
    ("ms", "Malayca", "Malezya"),
    ("bn", "Bengalce", "Bangladeş"),
    ("fa", "Farsça", "İran"),
]

for row_idx, (code, name, countries) in enumerate(languages, 2):
    ws3.cell(row=row_idx, column=1, value=code).border = thin_border
    ws3.cell(row=row_idx, column=2, value=name).border = thin_border
    ws3.cell(row=row_idx, column=3, value=countries).border = thin_border

ws3.freeze_panes = 'A2'

# Kaydet
output_path = "/Users/mustafayildiz/Documents/IW_Developments/scripts/kitapcik_toplu_yukleme_sablonu.xlsx"
wb.save(output_path)
print(f"✅ Şablon oluşturuldu: {output_path}")
